import openpyxl
from robot.libraries.BuiltIn import BuiltIn
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import requests
import win32com.client as win32
import os
from selenium.webdriver.common.action_chains import ActionChains
import pyautogui
import allure
import subprocess

class prerequisite(object):

        def __init__(self):
                pass
        @property
        def _sel_lib(self):
            return BuiltIn().get_library_instance('SeleniumLibrary')

        @property
        def _driver(self):
            return self._sel_lib.driver

        def get_data_values(self, file_path, sheet_name):
            wb = openpyxl.load_workbook(file_path)
            sheet = wb[sheet_name]
            data_values = {}
            # Get headers from the first row
            headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
            for row in range(2, sheet.max_row + 1):
                # Initialize dictionary for current row
                row_data = {}
                for col, header in enumerate(headers, start=1):
                    # Get value from current cell
                    value = sheet.cell(row=row, column=col).value
                    if isinstance(value, list):
                        value = ', '.join(map(str, value))  # Convert list to string
                    # Add value to row_data dictionary with key as column name
                    row_data[header] = value
                # Use data from the first column as key in data_values dictionary
                key = row_data[headers[0]]
                # If the key is not already in the dictionary, add it
                if key not in data_values:
                    data_values[key] = {header: [] for header in headers[1:]}
                # Add values for the current row
                for header, value in row_data.items():
                    if header != headers[0]:
                        data_values[key][header].append(value)
            result = [{key: value[0] for key, value in row_data.items()} for key, row_data in data_values.items()]

            split_data = [{key: value for key, value in item.items()} for item in result]
            # Remove keys with None values from each dictionary
            dict_list = []
            for item in split_data:
                item = {k: v for k, v in item.items() if v is not None}
                dict_list.append(item)
            return dict_list

