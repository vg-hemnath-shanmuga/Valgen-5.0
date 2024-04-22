import openpyxl
from robot.libraries.BuiltIn import BuiltIn
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import requests
import win32com.client as win32
import os
from selenium.webdriver.common.action_chains import ActionChains
import pyautogui
import allure
import subprocess

class CustomLibrary(object):

        def __init__(self):
                pass
        @property
        def _sel_lib(self):
            return BuiltIn().get_library_instance('SeleniumLibrary')

        @property
        def _driver(self):
            return self._sel_lib.driver

        def open_chrome_browser(self,url):
            """Return the True if Chrome browser opened """
            selenium = BuiltIn().get_library_instance('SeleniumLibrary')
            try:
                options = webdriver.ChromeOptions()
                options.add_argument('--disable-gpu')
                options.add_argument("disable-extensions")
                options.add_argument('--ignore-ssl-errors=yes')
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--use-fake-ui-for-media-stream')
                options.add_experimental_option('prefs', {
                    'credentials_enable_service': False,
                    'profile': {
                        'password_manager_enabled': False
                    }
                })
                options.add_experimental_option("excludeSwitches",["enable-automation","load-extension"])
                selenium.create_webdriver('Chrome',chrome_options=options)
                selenium.go_to(url)
                return True
            except:
                return False
            
        def convert_xls_2_xlsx(self, xls_path, xlsx_path):
            # Create temp xlsx-File
            if os.path.exists(xlsx_path): os.remove(xlsx_path)
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = 0
            wb = excel.Workbooks.Open(xls_path)
            wb.SaveAs(xlsx_path, FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
            wb.Close()

        def get_summary_view_details_from_excel(self, filepath):
            # get the summary vew details from excel file
            workbook = openpyxl.load_workbook(filepath)
            sheet_names = workbook.sheetnames
            # Choose the sheet by index (e.g., the first sheet)
            selected_sheet = workbook[sheet_names[0]]
            key1 = selected_sheet['A4'].value
            key2 = selected_sheet['B4'].value
            key3 = selected_sheet['C4'].value
            key4 = selected_sheet['D4'].value
            key5 = selected_sheet['E4'].value
            key6 = selected_sheet['F4'].value
            # # Accessing values using row and column indices
            summary_keys = [str(key1), str(key2), str(key3), str(key4), str(key5), str(key6)]
            # Assuming the row index is 4 (0-based) since you are trying to access row 5
            value1 = selected_sheet['A5'].value
            value2 = selected_sheet['B5'].value
            value3 = selected_sheet['C5'].value
            value4 = selected_sheet['D5'].value
            value5 = selected_sheet['E5'].value
            value6 = selected_sheet['F5'].value

            summary_values = [int(value1), int(value2), int(value3), int(value4), int(value5), int(value6)]
            summary = self.create_dictionary_from_two_lists(summary_keys,summary_values)
            summary = {x.replace(' ', ''): v 
                    for x, v in summary.items()}
            return  summary

        def get_project_ids_from_excel(self, filepath, projectids_count):
            # get the project ids from the Excel file
            workbook = openpyxl.load_workbook(filepath)
            # Assuming you want to read values from the first sheet
            sheet = workbook.worksheets[0]
            # Assuming A, B, C, D, E, F are column indices (1-based in openpyxl)
            col_index = 1
            # Assuming the row index is 14 (1-based)
            row_index = 14
            row_index_range = row_index + int(projectids_count)
            project_ids = []
            # Accessing values using row and column indices
            for row_No in range(row_index, row_index_range):
                project_id = sheet.cell(row=row_No, column=col_index).value
                # You can add additional processing here if needed
                # project_id = project_id.replace(' ', '')
                project_ids.append(str(project_id))

            return project_ids

        def create_dictionary_from_two_lists(self,key_list,value_list):
            # using dict() and zip() to convert lists to dictionary
            res = dict(zip(key_list, value_list))
            return  res

        def open_file(self, path):
            os.system(path)

        def print_screen(self):
            pyautogui.keyDown("printscreen")
            pyautogui.keyUp("printscreen")
            time.sleep(2)

        def open_headless_chrome(self,url):
            """Return the True if Chrome browser opened """
            selenium = BuiltIn().get_library_instance('SeleniumLibrary')
            options = webdriver.ChromeOptions()
            options.add_argument("--window-size=1440,900")
            options.add_argument('--disable-gpu')
            options.add_argument("disable-extensions")
            options.add_argument("--headless")
            options.add_experimental_option('prefs', {
                'credentials_enable_service': False,
                'profile': {
                    'password_manager_enabled': False
                }
            })
            options.add_experimental_option("excludeSwitches",["enable-automation","load-extension"])
            selenium.create_webdriver('Chrome',chrome_options=options)
            selenium.go_to(url)
        
        def click_using_javascript(self,locator):
            element = self._sel_lib.get_webelement(locator)
            self._driver.execute_script("arguments[0].click();", element)
        
        def javascript_click(self, locator):
            try:
                element = self._sel_lib.get_webelement(locator)
                self._driver.execute_script("arguments[0].click();", element)
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self._driver.save_screenshot("selenium-screenshot-"+filename + ".png")
            
        def get_text_by_using_javascript(self, locator):
            element = self._sel_lib.get_webelement(locator) 
            return self._driver.execute_script("return arguments[0].textContent;", element)

        def wait_until_time(self,arg):
                time.sleep(int(arg))
            
        def wait_until_element_clickable(self,locator):
            """ An Expectation for checking that an element is either invisible or not present on the DOM."""
            if locator.startswith("//") or locator.startswith("(//"):
               WebDriverWait(self._driver, 60).until(EC.element_to_be_clickable((By.XPATH, locator)))
            else:
               WebDriverWait(self._driver, 60).until(EC.element_to_be_clickable((By.ID, locator)))
        
        # def get_ms_excel_row_values_into_dictionary_based_on_key(self,filepath,keyName,sheetName=None):
        #     """Returns the dictionary of values given row in the MS Excel file """
            # workbook = xlrd.open_workbook(filepath)
            # snames = workbook.sheet_names()
            # dictVar = {}
            # if sheetName == None:
            #     sheetName = snames[0]      
            # if self.Verify_the_sheet_in_ms_excel_file(filepath,sheetName) == False:
            #     return dictVar
            # worksheet = workbook.sheet_by_name(sheetName)
            # noofrows = worksheet.nrows
            # dictVar = {}
            # headersList = worksheet.row_values(int(0))
            # for rowNo in range(1,int(noofrows)):
            #     rowValues = worksheet.row_values(int(rowNo))
            #     if str(rowValues[0])!= str(keyName):
            #         continue
            #     for rowIndex in range(0,len(rowValues)):
            #         cell_data = rowValues[rowIndex]
            #         if(str(cell_data) == "" or str(cell_data) == None):
            #             continue                    
            #         cell_data = self.get_unique_test_data(cell_data)
                
            #         dictVar[str(headersList[rowIndex])] = str(cell_data)
            # return dictVar 

        def get_ms_excel_row_values_into_dictionary_based_on_key(self, filepath, keyName, sheetName=None):
            """Returns the dictionary of values given row in the MS Excel file"""
            workbook = openpyxl.load_workbook(filepath)
            snames = workbook.sheetnames
            dictVar = {}

            if sheetName is None:
                sheetName = snames[0]

            if sheetName not in snames or not self.Verify_the_sheet_in_ms_excel_file(filepath, sheetName):
                return dictVar

            worksheet = workbook[sheetName]
            headersList = [str(cell.value) for cell in worksheet[1]]

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if str(row[0]) != str(keyName):
                    continue

                for rowIndex, cell_data in enumerate(row):
                    if cell_data is None or cell_data == "":
                        continue

                    # Ensure that get_unique_test_data is implemented correctly
                    cell_data = self.get_unique_test_data(cell_data)

                    dictVar[str(headersList[rowIndex])] = str(cell_data)

            return dictVar                 

        def get_unique_test_data(self,testdata):
            """Returns the unique if data contains unique word """
            ts = time.strftime("%H%M%S")
            unique_string = str(ts)
            testdata = testdata.replace("UNIQUE",unique_string)
            testdata = testdata.replace("Unique",unique_string)
            testdata = testdata.replace("unique",unique_string)
            return testdata

        def Verify_the_sheet_in_ms_excel_file(self,filepath,sheetName):
            """Returns the True if the specified work sheets exist in the specifed MS Excel file else False"""
            # workbook = xlrd.open_workbook(filepath)
            # snames = workbook.sheet_names()
            workbook = openpyxl.load_workbook(filepath)
            snames = workbook.sheetnames
            sStatus = False        
            if sheetName == None:
                return True
            else:
                for sname in snames:
                    if sname.lower() == sheetName.lower():
                        wsname = sname
                        sStatus = True
                        break
                if sStatus == False:
                    print ("Error: The specified sheet: "+str(sheetName)+" doesn't exist in the specified file: " +str(filepath))
            return sStatus
        
        def clear_text_field(self, locator):
            element = self._sel_lib.get_webelement(locator)
            self._driver.execute_script('arguments[0].value = "";', element)

        def javascript_input_text(self,locator, text):
            try:
                element = self._sel_lib.get_webelement(locator)
                self._driver.execute_script('arguments[0].value = arguments[1];', element, text)
                self._driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", element)
                self._driver.execute_script('arguments[0].focus();', element)
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self._driver.save_screenshot("selenium-screenshot-"+filename + ".png")

        def download_PDF(self, project_folder, filename):
            pdf_url = self._driver.current_url
            response = requests.get(pdf_url)
            file_name = os.path.join(project_folder, filename)
            with open(file_name, 'wb') as f:
                f.write(response.content)
        
        def click_calendar(self, locator):
            element = self._sel_lib.get_webelement(locator)
            # ActionChains(self._driver).move_by_offset(578, 465).click().perform()
            ActionChains(self._driver).move_to_element_with_offset(element, 95, 2).click().perform()

        def select_canvas_checkbox(self, locator):
            element = self._sel_lib.get_webelement(locator)
            time.sleep(1)
            ActionChains(self._driver).click(element).key_down(Keys.CONTROL).send_keys(Keys.ARROW_LEFT).key_up(Keys.CONTROL).perform()
            time.sleep(1)
            ActionChains(self._driver).click(element).key_down(Keys.SPACE).key_up(Keys.SPACE).perform()

        def click_calendar_icon_in_vlms(self, locator):
            element = self._sel_lib.get_webelement(locator)
            # ActionChains(self._driver).move_by_offset(578, 465).click().perform()
            ActionChains(self._driver).move_to_element_with_offset(element, 61, 2).click().perform()

        def click_element_with_offset(self, locator, x, y):
            element = self._sel_lib.get_webelement(locator)
            ActionChains(self._driver).move_to_element_with_offset(element, x, y).click().perform()

        def edit_document_in_vlms(self):
            # ActionChains(self._driver).move_by_offset(800, 450).click().perform()
            ActionChains(self._driver).send_keys("Review It").perform()
            ActionChains(self._driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
            # with pyautogui.hold('shift'):
            # pyautogui.press(['R', 'E', 'V', 'I','E','W',' ','I','T'])
            # pyautogui.press('enter')
            time.sleep(3)
        
        def click_in_document(self,x,y):
            ActionChains(self._driver).move_by_offset(x, y).click().perform()
        
        def scroll_down_in_document(self, limit):
            """ moving slder to some extent along x-axis """
            for _ in range(int(limit)):
                ActionChains(self._driver).key_down(Keys.PAGE_DOWN).key_up(Keys.PAGE_DOWN).perform()
                # pyautogui.PAUSE = 0.1
                # pyautogui.FAILSAFE = False
                # pyautogui.press('enter')
                # for _ in range(int(notches)):
                #     pyautogui.press('down')
            time.sleep(3)
        
        def switch_to_parent_frame(self):
            self._driver.switch_to.parent_frame()

        def screenshot_page(self,png_name):
            ul = BuiltIn().get_library_instance('SeleniumLibrary')
            path = ul.capture_page_screenshot()
            allure.attach.file(path, name=png_name, attachment_type=allure.attachment_type.JPG)
            return path

        def upload_supporting_documents(self, choosebutton, filepath):
                element = self._driver.find_element("xpath", choosebutton)
                self._driver.execute_script("arguments[0].setAttribute('style', 'top: 0px;');",element)
                time.sleep(2)
                element = self._driver.find_element("xpath", choosebutton)
                element.send_keys(filepath)

        def open_file_and_take_screenshot(self, path, file_name):
            subprocess.Popen([path], shell=True)
            time.sleep(7)
            screenshot = pyautogui.screenshot()
            screenshot.save(file_name)
            # if kill!='None': os.system("taskkill /f /im "+ app +".exe")

        def input_text_with_offset(self, locator, x, y, text):
            element = self._sel_lib.get_webelement(locator)
            time.sleep(3)
            ActionChains(self._driver).move_to_element_with_offset(element, x, y).click().send_keys(text).perform()

        def double_click_element_with_offset(self, locator, x, y):
            element = self._sel_lib.get_webelement(locator)
            ActionChains(self._driver).move_to_element_with_offset(element, x, y).double_click().perform()

        # def get_data_values(self, file_path, sheet_name):
        #     wb = openpyxl.load_workbook(file_path)
        #     sheet = wb[sheet_name]
        #     data_values = {}
        #     # Get headers from the first row
        #     headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
        #     for row in range(2, sheet.max_row + 1):
        #         # Initialize dictionary for current row
        #         row_data = {}
        #         for col, header in enumerate(headers, start=1):
        #             # Get value from current cell
        #             value = sheet.cell(row=row, column=col).value
        #             if isinstance(value, list):
        #                 value = ', '.join(map(str, value))  # Convert list to string
        #             # Add value to row_data dictionary with key as column name
        #             row_data[header] = value
        #         # Use data from the first column as key in data_values dictionary
        #         key = row_data[headers[0]]
        #         # If the key is not already in the dictionary, add it
        #         if key not in data_values:
        #             data_values[key] = {header: [] for header in headers[1:]}
        #         # Add values for the current row
        #         for header, value in row_data.items():
        #             if header != headers[0]:
        #                 data_values[key][header].append(value)
        #     result = [{key: value[0] for key, value in row_data.items()} for key, row_data in data_values.items()]
        #     split_data = [{key: value for key, value in item.items()} for item in result]
        #     # Remove keys with None values from each dictionary
        #     dict_list = []
        #     for item in split_data:
        #         item = {k: v for k, v in item.items() if v is not None}
        #         dict_list.append(item)
        #     return dict_list